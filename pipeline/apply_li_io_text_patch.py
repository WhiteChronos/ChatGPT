#!/usr/bin/env python3
"""Aplica alterações textuais controladas em LI de Entradas e Saídas.

O script altera somente o valor de células existentes, atualiza paginação quando
solicitado e bloqueia a saída se geometria, estilos, fontes, bordas, mesclagens,
dimensões, impressão ou objetos divergirem do arquivo-fonte.

Exemplo de patch JSON:
{
  "changes": [
    {"sheet": "2", "cell": "H30", "value": "SINAL DE INCÊNDIO"},
    {"sheet": "3", "cell": "Y24", "formula": "=I24+Q24"}
  ],
  "update_pagination": true
}
"""

from __future__ import annotations

import argparse
import json
import shutil
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_to_tuple

from pipeline.li_io_standard import compare_layout, findings_to_json, validate_standard


@dataclass(frozen=True)
class CellChange:
    sheet: str
    cell: str
    old_value: Any
    new_value: Any
    style_id_before: int
    style_id_after: int


class PatchError(ValueError):
    """Erro de contrato do patch textual."""


def _is_xlsm(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def _load(path: Path):
    return load_workbook(path, data_only=False, keep_links=True, keep_vba=_is_xlsm(path))


def _validate_patch(patch: dict[str, Any]) -> list[dict[str, Any]]:
    changes = patch.get("changes")
    if not isinstance(changes, list) or not changes:
        raise PatchError("O patch deve conter uma lista não vazia em 'changes'.")

    seen: set[tuple[str, str]] = set()
    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(changes, start=1):
        if not isinstance(item, dict):
            raise PatchError(f"Alteração {index} não é um objeto.")
        sheet = str(item.get("sheet", "")).strip()
        cell = str(item.get("cell", "")).strip().upper()
        has_value = "value" in item
        has_formula = "formula" in item
        if not sheet or not cell or has_value == has_formula:
            raise PatchError(
                f"Alteração {index} deve ter sheet, cell e exatamente um entre value/formula."
            )
        try:
            coordinate_to_tuple(cell)
        except ValueError as exc:
            raise PatchError(f"Coordenada inválida: {cell}") from exc
        key = (sheet, cell)
        if key in seen:
            raise PatchError(f"Célula repetida no patch: {sheet}!{cell}")
        seen.add(key)
        new_value = item.get("formula") if has_formula else item.get("value")
        if has_formula and (not isinstance(new_value, str) or not new_value.startswith("=")):
            raise PatchError(f"Fórmula inválida em {sheet}!{cell}")
        normalized.append({"sheet": sheet, "cell": cell, "new_value": new_value})
    return normalized


def _pagination_cells(index: int) -> tuple[str, str]:
    return ("N2", "P2") if index == 1 else ("AQ2", "AS2")


def apply_patch(
    source: str | Path,
    patch_path: str | Path,
    output: str | Path,
    *,
    report_path: str | Path | None = None,
) -> dict[str, Any]:
    source_path = Path(source)
    output_path = Path(output)
    patch = json.loads(Path(patch_path).read_text(encoding="utf-8"))
    normalized = _validate_patch(patch)

    if output_path.resolve() == source_path.resolve():
        raise PatchError("A saída não pode sobrescrever o arquivo-fonte.")
    if output_path.suffix.lower() != source_path.suffix.lower():
        raise PatchError("A extensão da saída deve ser igual à do arquivo-fonte.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)
    workbook = _load(output_path)
    cell_changes: list[CellChange] = []

    try:
        for item in normalized:
            sheet_name = item["sheet"]
            coordinate = item["cell"]
            if sheet_name not in workbook.sheetnames:
                raise PatchError(f"Aba inexistente: {sheet_name}")
            worksheet = workbook[sheet_name]
            row, column = coordinate_to_tuple(coordinate)
            if row > worksheet.max_row or column > worksheet.max_column:
                raise PatchError(
                    f"Célula fora da matriz existente: {sheet_name}!{coordinate}"
                )
            cell = worksheet[coordinate]
            if isinstance(cell, MergedCell):
                raise PatchError(
                    f"{sheet_name}!{coordinate} não é a célula âncora da mesclagem."
                )
            before_style = cell.style_id
            old_value = cell.value
            cell.value = item["new_value"]
            cell_changes.append(
                CellChange(
                    sheet=sheet_name,
                    cell=coordinate,
                    old_value=old_value,
                    new_value=item["new_value"],
                    style_id_before=before_style,
                    style_id_after=cell.style_id,
                )
            )

        if patch.get("update_pagination", False):
            total = len(workbook.sheetnames)
            for index, sheet_name in enumerate(workbook.sheetnames, start=1):
                worksheet = workbook[sheet_name]
                current_cell, total_cell = _pagination_cells(index)
                for coordinate, value in ((current_cell, index), (total_cell, total)):
                    row, column = coordinate_to_tuple(coordinate)
                    if row > worksheet.max_row or column > worksheet.max_column:
                        raise PatchError(
                            f"Célula de paginação ausente: {sheet_name}!{coordinate}"
                        )
                    cell = worksheet[coordinate]
                    if isinstance(cell, MergedCell):
                        raise PatchError(
                            f"Célula de paginação não é âncora: {sheet_name}!{coordinate}"
                        )
                    if cell.value != value:
                        before_style = cell.style_id
                        old_value = cell.value
                        cell.value = value
                        cell_changes.append(
                            CellChange(
                                sheet=sheet_name,
                                cell=coordinate,
                                old_value=old_value,
                                new_value=value,
                                style_id_before=before_style,
                                style_id_after=cell.style_id,
                            )
                        )

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output_path)
    except Exception:
        output_path.unlink(missing_ok=True)
        raise

    layout_findings = compare_layout(source_path, output_path)
    standard_findings = validate_standard(output_path)
    style_findings = [
        {
            "code": "LI-CELL-STYLE-MUTATION",
            "severity": "CRITICAL",
            "message": "O estilo da célula foi alterado durante a aplicação textual.",
            "sheet": item.sheet,
            "cell": item.cell,
        }
        for item in cell_changes
        if item.style_id_before != item.style_id_after
    ]

    findings = findings_to_json(layout_findings + standard_findings) + style_findings
    critical = [item for item in findings if item["severity"] == "CRITICAL"]
    report = {
        "source": str(source_path),
        "output": str(output_path),
        "mode": "TEXT_ONLY",
        "update_pagination": bool(patch.get("update_pagination", False)),
        "changed_cells": [asdict(item) for item in cell_changes],
        "findings": findings,
        "critical_count": len(critical),
        "approved_for_emission": not critical,
    }

    if report_path:
        Path(report_path).write_text(
            json.dumps(report, ensure_ascii=False, indent=2, default=str) + "\n",
            encoding="utf-8",
        )

    if critical:
        output_path.unlink(missing_ok=True)
        raise PatchError(
            f"Saída bloqueada: {len(critical)} finding(s) crítico(s)."
        )
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source")
    parser.add_argument("patch")
    parser.add_argument("output")
    parser.add_argument("--report")
    args = parser.parse_args(argv)

    try:
        report = apply_patch(
            args.source,
            args.patch,
            args.output,
            report_path=args.report,
        )
    except (PatchError, json.JSONDecodeError) as exc:
        print(f"CRITICAL LI-TEXT-PATCH: {exc}")
        return 1

    print(
        f"changed_cells={len(report['changed_cells'])} "
        f"critical={report['critical_count']} approved={report['approved_for_emission']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

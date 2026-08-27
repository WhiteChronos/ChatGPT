#!/usr/bin/env python3
"""Governança determinística para LI de Entradas e Saídas.

O módulo implementa a regra principal do padrão:
- conteúdo pertence ao documento de destino;
- layout pertence ao modelo/arquivo-fonte e é imutável;
- quantidade de abas e paginação pertencem ao documento de destino;
- somente texto, valores, fórmulas autorizadas e paginação podem mudar.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl é obrigatório para validar LI de Entradas e Saídas") from exc


@dataclass(frozen=True)
class Finding:
    code: str
    severity: str
    message: str
    sheet: str | None = None
    cell: str | None = None


def sha256_file(path: str | Path) -> str:
    p = Path(path)
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _round(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 12)
    return value


def _style_hash(ws: Any) -> str:
    payload = [(cell.coordinate, cell.style_id) for row in ws.iter_rows() for cell in row]
    encoded = json.dumps(payload, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _image_signature(image: Any) -> dict[str, Any]:
    anchor = getattr(image, "anchor", None)
    signature: dict[str, Any] = {
        "width": _round(getattr(image, "width", None)),
        "height": _round(getattr(image, "height", None)),
        "anchor_type": type(anchor).__name__ if anchor is not None else None,
    }
    if anchor is not None:
        for name in ("_from", "to"):
            point = getattr(anchor, name, None)
            if point is not None:
                signature[name] = {
                    "col": getattr(point, "col", None),
                    "row": getattr(point, "row", None),
                    "colOff": getattr(point, "colOff", None),
                    "rowOff": getattr(point, "rowOff", None),
                }
    return signature


def sheet_layout_signature(ws: Any) -> dict[str, Any]:
    """Retorna assinatura estrutural sem valores de células."""
    return {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_ranges": sorted(str(rng) for rng in ws.merged_cells.ranges),
        "row_heights": {
            str(index): _round(ws.row_dimensions[index].height)
            for index in range(1, ws.max_row + 1)
            if ws.row_dimensions[index].height is not None
        },
        "column_widths": {
            key: _round(dimension.width)
            for key, dimension in sorted(ws.column_dimensions.items())
            if dimension.width is not None
        },
        "row_hidden": {
            str(index): bool(ws.row_dimensions[index].hidden)
            for index in range(1, ws.max_row + 1)
            if ws.row_dimensions[index].hidden
        },
        "column_hidden": {
            key: bool(dimension.hidden)
            for key, dimension in sorted(ws.column_dimensions.items())
            if dimension.hidden
        },
        "print_area": str(ws.print_area),
        "print_title_rows": str(ws.print_title_rows),
        "print_title_cols": str(ws.print_title_cols),
        "page_setup": {
            "orientation": ws.page_setup.orientation,
            "paper_size": ws.page_setup.paperSize,
            "scale": ws.page_setup.scale,
            "fit_to_width": ws.page_setup.fitToWidth,
            "fit_to_height": ws.page_setup.fitToHeight,
            "horizontal_centered": ws.print_options.horizontalCentered,
            "vertical_centered": ws.print_options.verticalCentered,
        },
        "page_margins": {
            "left": _round(ws.page_margins.left),
            "right": _round(ws.page_margins.right),
            "top": _round(ws.page_margins.top),
            "bottom": _round(ws.page_margins.bottom),
            "header": _round(ws.page_margins.header),
            "footer": _round(ws.page_margins.footer),
        },
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "sheet_view": [
            {
                "showGridLines": view.showGridLines,
                "zoomScale": view.zoomScale,
                "zoomScaleNormal": view.zoomScaleNormal,
            }
            for view in ws.views.sheetView
        ],
        "style_ids_hash": _style_hash(ws),
        "images": [_image_signature(image) for image in getattr(ws, "_images", [])],
        "chart_count": len(getattr(ws, "_charts", [])),
        "data_validation_count": len(getattr(ws.data_validations, "dataValidation", [])),
        "conditional_formatting_count": len(ws.conditional_formatting),
    }


def workbook_layout_signature(path: str | Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, keep_links=True)
    return {
        "sheet_order": list(wb.sheetnames),
        "sheets": {title: sheet_layout_signature(wb[title]) for title in wb.sheetnames},
        "defined_names": sorted(name.name for name in wb.defined_names.values()),
    }


def compare_layout(
    reference_path: str | Path,
    candidate_path: str | Path,
    *,
    allow_sheet_count_change: bool = False,
) -> list[Finding]:
    """Compara layout sem considerar conteúdo textual.

    Em revisão de documento existente, `allow_sheet_count_change` deve permanecer falso.
    Em elaboração de novo documento, a quantidade de abas pode variar, mas cada aba
    precisa usar um arquétipo estrutural já aprovado no modelo de referência.
    """
    reference = workbook_layout_signature(reference_path)
    candidate = workbook_layout_signature(candidate_path)
    findings: list[Finding] = []

    ref_order = reference["sheet_order"]
    cand_order = candidate["sheet_order"]

    if not allow_sheet_count_change and ref_order != cand_order:
        findings.append(
            Finding(
                "LI-LAYOUT-SHEETS",
                "CRITICAL",
                f"Abas alteradas. Referência={ref_order!r}; candidato={cand_order!r}",
            )
        )
    elif allow_sheet_count_change:
        if not cand_order:
            findings.append(Finding("LI-SHEETS-EMPTY", "CRITICAL", "Documento sem abas"))
        elif not ref_order:
            findings.append(Finding("LI-REFERENCE-EMPTY", "CRITICAL", "Modelo sem abas"))
        else:
            ref_cover = reference["sheets"][ref_order[0]]
            if candidate["sheets"][cand_order[0]] != ref_cover:
                findings.append(
                    Finding("LI-COVER-ARCHETYPE", "CRITICAL", "A capa não corresponde ao arquétipo aprovado", cand_order[0])
                )
            archetypes = list(reference["sheets"].values())[1:]
            for title in cand_order[1:]:
                if candidate["sheets"][title] not in archetypes:
                    findings.append(
                        Finding(
                            "LI-SHEET-ARCHETYPE",
                            "CRITICAL",
                            "Folha não corresponde a nenhum arquétipo de layout aprovado",
                            title,
                        )
                    )

    for title in set(ref_order) & set(cand_order):
        if reference["sheets"][title] != candidate["sheets"][title]:
            ref_sheet = reference["sheets"][title]
            cand_sheet = candidate["sheets"][title]
            changed = sorted(key for key in ref_sheet if ref_sheet.get(key) != cand_sheet.get(key))
            findings.append(
                Finding(
                    "LI-LAYOUT-MUTATION",
                    "CRITICAL",
                    "Alteração estrutural detectada: " + ", ".join(changed),
                    title,
                )
            )

    return findings


def _as_int(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def validate_pagination(path: str | Path) -> list[Finding]:
    wb = load_workbook(path, data_only=False, keep_links=True)
    findings: list[Finding] = []
    total = len(wb.sheetnames)

    for index, title in enumerate(wb.sheetnames, start=1):
        ws = wb[title]
        if index == 1:
            current_cell, total_cell = "N2", "P2"
        else:
            current_cell, total_cell = "AQ2", "AS2"

        current = _as_int(ws[current_cell].value)
        declared_total = _as_int(ws[total_cell].value)
        if current != index:
            findings.append(
                Finding(
                    "LI-PAGE-CURRENT",
                    "CRITICAL",
                    f"Folha declarada {current!r}; esperado {index}",
                    title,
                    current_cell,
                )
            )
        if declared_total != total:
            findings.append(
                Finding(
                    "LI-PAGE-TOTAL",
                    "CRITICAL",
                    f"Total declarado {declared_total!r}; esperado {total}",
                    title,
                    total_cell,
                )
            )

    return findings


def validate_note_placement(path: str | Path) -> list[Finding]:
    """Bloqueia nota genérica inserida como linha de mapa de memória."""
    wb = load_workbook(path, data_only=False, keep_links=True)
    findings: list[Finding] = []
    for title in wb.sheetnames:
        ws = wb[title]
        header = " ".join(str(ws.cell(5, col).value or "") for col in range(1, min(ws.max_column, 50) + 1)).upper()
        is_memory_map = "TAG" in header and ("TIPO DE SINAL" in header or "INTERFACE" in header)
        if not is_memory_map:
            continue
        for row in range(6, ws.max_row + 1):
            first = str(ws.cell(row, 1).value or "").strip().upper()
            if first in {"NOTA", "NOTE", "OBSERVAÇÃO", "OBSERVACAO"}:
                findings.append(
                    Finding(
                        "LI-NOTE-PLACEMENT",
                        "CRITICAL",
                        "Nota genérica inserida na folha de mapa de memória; mover para a folha de notas",
                        title,
                        f"A{row}",
                    )
                )
    return findings


def validate_full_document_codes(path: str | Path) -> list[Finding]:
    wb = load_workbook(path, data_only=False, keep_links=True)
    findings: list[Finding] = []
    prefixes = ("DE-", "MD-", "ET-", "FD-", "LI-")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                text = value.strip().upper()
                if not text.startswith(prefixes):
                    continue
                code = text.split()[0].rstrip(";,:.")
                if code.count("-") < 4 or "RPJ" not in code:
                    findings.append(
                        Finding(
                            "LI-REFERENCE-TRUNCATED",
                            "HIGH",
                            f"Referência possivelmente truncada: {value!r}",
                            ws.title,
                            cell.coordinate,
                        )
                    )
    return findings


def validate_standard(path: str | Path) -> list[Finding]:
    findings: list[Finding] = []
    findings.extend(validate_pagination(path))
    findings.extend(validate_note_placement(path))
    findings.extend(validate_full_document_codes(path))
    return findings


def findings_to_json(findings: Iterable[Finding]) -> list[dict[str, Any]]:
    return [asdict(item) for item in findings]


def _print_findings(findings: list[Finding]) -> None:
    for item in findings:
        location = ""
        if item.sheet:
            location += f" [{item.sheet}"
            if item.cell:
                location += f"!{item.cell}"
            location += "]"
        print(f"{item.severity} {item.code}{location}: {item.message}")
    critical = sum(item.severity == "CRITICAL" for item in findings)
    print(f"findings={len(findings)} critical={critical}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    p_compare = sub.add_parser("compare", help="Comparar layout do original e revisado")
    p_compare.add_argument("reference")
    p_compare.add_argument("candidate")
    p_compare.add_argument("--allow-sheet-count-change", action="store_true")
    p_compare.add_argument("--report")

    p_validate = sub.add_parser("validate-standard", help="Validar paginação e regras da LI E/S")
    p_validate.add_argument("workbook")
    p_validate.add_argument("--report")

    p_fingerprint = sub.add_parser("fingerprint", help="Gerar assinatura estrutural")
    p_fingerprint.add_argument("workbook")
    p_fingerprint.add_argument("--out")

    args = parser.parse_args(argv)

    if args.command == "compare":
        findings = compare_layout(
            args.reference,
            args.candidate,
            allow_sheet_count_change=args.allow_sheet_count_change,
        )
        findings.extend(validate_standard(args.candidate))
        if args.report:
            Path(args.report).write_text(json.dumps(findings_to_json(findings), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        _print_findings(findings)
        return 1 if any(item.severity == "CRITICAL" for item in findings) else 0

    if args.command == "validate-standard":
        findings = validate_standard(args.workbook)
        if args.report:
            Path(args.report).write_text(json.dumps(findings_to_json(findings), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        _print_findings(findings)
        return 1 if any(item.severity == "CRITICAL" for item in findings) else 0

    signature = workbook_layout_signature(args.workbook)
    content = json.dumps(signature, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
    if args.out:
        Path(args.out).write_text(content, encoding="utf-8")
    else:
        print(content, end="")
    return 0


LI_IO_MODEL_ID = "LI_IO_PETROBRAS_AUTOMACAO_V1_0"
FULL_DOCUMENT_CODE_RE = re.compile(
    r"^(?:MD|ET|LI|FD|DE)-\d{4}\.\d{2}-\d{4}-\d{3}-RPJ-\d{4}(?:=[0-9A-Z]+)?$"
)
MANIFEST_ALLOWED_CHANGES = {
    "TEXT",
    "TECHNICAL_VALUES",
    "TEXTUAL_QUANTITIES",
    "AUTHORIZED_FORMULAS",
    "PAGE_NUMBER",
    "TOTAL_PAGES",
}


@dataclass(frozen=True)
class ManifestFinding:
    code: str
    severity: str
    message: str
    item_id: str | None = None


def validate_manifest_io_summary(summary: dict[str, Any], item_id: str | None = None) -> list[ManifestFinding]:
    findings: list[ManifestFinding] = []
    for signal_type, counts in summary.items():
        try:
            used = int(counts["used"])
            reserve = int(counts["reserve"])
            total = int(counts["total"])
        except (KeyError, TypeError, ValueError):
            findings.append(ManifestFinding("LI-IO-COUNT-FORMAT", "CRITICAL", f"Resumo inválido para {signal_type}", item_id))
            continue
        if used + reserve != total:
            findings.append(
                ManifestFinding(
                    "LI-IO-COUNT-MISMATCH",
                    "CRITICAL",
                    f"{signal_type}: uso ({used}) + reserva ({reserve}) != total ({total})",
                    item_id,
                )
            )
    return findings


def validate_manifest_reference_codes(codes: Iterable[str], item_id: str | None = None) -> list[ManifestFinding]:
    findings: list[ManifestFinding] = []
    for code in codes:
        if not FULL_DOCUMENT_CODE_RE.match(str(code).strip()):
            findings.append(
                ManifestFinding(
                    "LI-IO-REFERENCE-CODE",
                    "HIGH",
                    f"Código documental incompleto ou inválido: {code!r}",
                    item_id,
                )
            )
    return findings


def validate_li_io_document(doc: dict[str, Any]) -> list[ManifestFinding]:
    """Valida o manifesto de emissão da LI E/S sem abrir o XLSX."""
    findings: list[ManifestFinding] = []
    did = str(doc.get("id", "UNKNOWN"))
    subtype = str(doc.get("document_subtype", "")).upper()
    if subtype not in {"LI_IO", "ENTRADAS_E_SAIDAS"}:
        return findings

    if doc.get("model_id") != LI_IO_MODEL_ID:
        findings.append(ManifestFinding("LI-IO-MODEL", "CRITICAL", f"Model ID deve ser {LI_IO_MODEL_ID}", did))
    if doc.get("sheet_count_source") != "TARGET_DOCUMENT":
        findings.append(ManifestFinding("LI-IO-SHEET-SOURCE", "CRITICAL", "Quantidade de abas deve ser definida pelo documento de destino", did))
    if doc.get("template_sheet_count_forced") is True:
        findings.append(ManifestFinding("LI-IO-TEMPLATE-SHEET-COUNT", "CRITICAL", "É proibido copiar a quantidade de abas do modelo visual", did))

    actual = doc.get("final_page_count")
    declared = doc.get("declared_total_pages")
    target = doc.get("target_sheet_count")
    if actual is None or target is None or actual != target:
        findings.append(ManifestFinding("LI-IO-TARGET-SHEET-COUNT", "CRITICAL", "Quantidade final deve corresponder à necessidade do destino", did))
    if declared != actual:
        findings.append(ManifestFinding("LI-IO-PAGINATION", "CRITICAL", "Paginação declarada não corresponde à quantidade final", did))

    checks = {
        "layout_signature_match": ("LI-IO-LAYOUT-SIGNATURE", "Assinatura estrutural do layout não confere"),
        "font_signature_match": ("LI-IO-FONT-SIGNATURE", "Fonte ou tamanho de fonte diverge do padrão"),
        "cell_dimensions_match": ("LI-IO-CELL-DIMENSIONS", "Dimensões de células/linhas/colunas divergem do padrão"),
        "print_settings_match": ("LI-IO-PRINT-SETTINGS", "Configuração de impressão diverge do padrão"),
        "only_authorized_cell_changes": ("LI-IO-TEXT-ONLY", "Foram detectadas alterações fora das células autorizadas"),
    }
    for field, (code, message) in checks.items():
        if doc.get(field) is not True:
            findings.append(ManifestFinding(code, "CRITICAL", message, did))

    if doc.get("notes_sheet_role") != "NOTES_SUMMARY_REFERENCES":
        findings.append(ManifestFinding("LI-IO-NOTES-SHEET", "CRITICAL", "Notas devem permanecer na folha de notas/resumo/referências", did))
    if doc.get("ad_hoc_note_on_final_sheet") is True:
        findings.append(ManifestFinding("LI-IO-FINAL-NOTE", "CRITICAL", "Nota genérica indevida na última folha/mapa de memória", did))

    changes = set(doc.get("changes", []))
    forbidden = changes - MANIFEST_ALLOWED_CHANGES
    if forbidden:
        findings.append(ManifestFinding("LI-IO-CHANGE-TYPE", "CRITICAL", f"Alterações não autorizadas: {', '.join(sorted(forbidden))}", did))

    for field, code in (
        ("model_fingerprint", "LI-IO-MODEL-FINGERPRINT"),
        ("source_fingerprint", "LI-IO-SOURCE-FINGERPRINT"),
        ("output_fingerprint", "LI-IO-OUTPUT-FINGERPRINT"),
    ):
        if not doc.get(field):
            findings.append(ManifestFinding(code, "CRITICAL", f"{field} não registrado", did))

    findings.extend(validate_manifest_io_summary(doc.get("io_summary", {}), did))
    findings.extend(validate_manifest_reference_codes(doc.get("reference_codes", []), did))
    return findings


if __name__ == "__main__":
    raise SystemExit(main())

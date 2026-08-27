"""Regressões do padrão definitivo de LI de Entradas e Saídas v4.6."""

import json
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

import pipeline.agents_v4_6 as agents
from pipeline.apply_li_io_text_patch import PatchError, apply_patch
from pipeline.audit_document_plugins import load_registry
from pipeline.validate_engineering import validate_project
from pipeline.xlsx_layout_guard import compare_layout


def valid_li(**overrides):
    doc = {
        "id": "LI-TEST",
        "document_type": "LI",
        "document_subtype": "LI_IO",
        "understanding_status": "APROVADO",
        "layout_immutable": True,
        "datasheet_layout_immutable": True,
        "reference_validation_required": True,
        "template_fingerprint": "sha256:template",
        "model_fingerprint": "sha256:model",
        "source_fingerprint": "sha256:source",
        "output_fingerprint": "sha256:output",
        "model_id": agents.LI_IO_MODEL_ID,
        "sheet_count_source": "TARGET_DOCUMENT",
        "template_sheet_count_forced": False,
        "target_sheet_count": 4,
        "final_page_count": 4,
        "declared_total_pages": 4,
        "layout_signature_match": True,
        "font_signature_match": True,
        "cell_dimensions_match": True,
        "print_settings_match": True,
        "only_authorized_cell_changes": True,
        "notes_sheet_role": "NOTES_SUMMARY_REFERENCES",
        "ad_hoc_note_on_final_sheet": False,
        "changes": ["TEXT", "TECHNICAL_VALUES", "AUTHORIZED_FORMULAS", "TOTAL_PAGES"],
        "io_summary": {
            "DI": {"used": 7, "reserve": 9, "total": 16},
            "DO": {"used": 4, "reserve": 12, "total": 16},
        },
        "reference_codes": [
            "DE-3501.01-8210-800-RPJ-1000",
            "DE-3501.01-8210-800-RPJ-1001",
            "MD-3501.01-8210-800-RPJ-1000",
        ],
    }
    doc.update(overrides)
    return doc


def codes(doc):
    return {finding.code for finding in validate_project({"documents": [doc]})}


def test_agents_and_golden_checks_are_registered():
    assert "LIInputOutputStandardAgent" in agents.AGENTS
    assert "LITextPatchAgent" in agents.AGENTS
    assert "DocumentToolingDiscoveryAgent" in agents.AGENTS
    assert agents.GOLDEN_CHECKS["sheet_count_target_driven"] is True
    assert agents.GOLDEN_CHECKS["font_size_immutable"] is True
    assert agents.GOLDEN_CHECKS["plugin_auto_install_forbidden"] is True


def test_valid_li_manifest_passes():
    assert validate_project({"documents": [valid_li()]}) == []


def test_li_manifest_rejects_template_sheet_count_and_layout_mutation():
    result = codes(valid_li(
        sheet_count_source="VISUAL_TEMPLATE",
        template_sheet_count_forced=True,
        font_signature_match=False,
    ))
    assert "LI-IO-SHEET-SOURCE" in result
    assert "LI-IO-TEMPLATE-SHEET-COUNT" in result
    assert "DOC-TEMPLATE-SHEET-COUNT" in result
    assert "LI-IO-FONT-SIGNATURE" in result


def test_li_manifest_rejects_wrong_total_and_final_note():
    result = codes(valid_li(
        io_summary={"DI": {"used": 7, "reserve": 9, "total": 17}},
        ad_hoc_note_on_final_sheet=True,
    ))
    assert "LI-IO-COUNT-MISMATCH" in result
    assert "LI-IO-FINAL-NOTE" in result


def test_registry_is_curated_and_never_auto_installs():
    registry = load_registry("plugins/document_tooling_registry.json")
    assert registry["policy"]["auto_install_forbidden"] is True
    assert all(item["auto_install"] is False for item in registry["repositories"])
    names = {entry["full_name"] for entry in registry["repositories"]}
    for name in (
        "python-openxml/python-docx",
        "LibreOffice/core",
        "docling-project/docling",
        "microsoft/markitdown",
        "dotnet/Open-XML-SDK",
        "mozman/ezdxf",
    ):
        assert name in names


def test_public_datacenter_manifest_is_redacted_and_target_driven():
    data = json.loads(Path("datacenter/LI_IO_STANDARD.json").read_text(encoding="utf-8"))
    assert data["status"] == "ACTIVE"
    assert data["sheet_count_policy"] == "TARGET_DOCUMENT_DRIVEN"
    assert data["canonical_sheet_count_is_not_target_constraint"] is True
    assert data["layout_lock"] == "TEXT_ONLY"
    assert data["binary_storage_policy"] == "PRIVATE_DATACENTER_ONLY"
    assert "canonical_model_source" not in data


def make_xlsx(path: Path, value: str, style_id: str = "1", width: str = "12") -> None:
    parts = {
        "[Content_Types].xml": "<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'/>",
        "xl/workbook.xml": "<workbook xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'><sheets><sheet name='A' sheetId='1'/></sheets></workbook>",
        "xl/styles.xml": "<styleSheet xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'><fonts count='1'><font><sz val='8'/><name val='Arial'/></font></fonts></styleSheet>",
        "xl/worksheets/sheet1.xml": (
            "<worksheet xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'>"
            f"<cols><col min='1' max='1' width='{width}' customWidth='1'/></cols>"
            f"<sheetData><row r='1' ht='15' customHeight='1'><c r='A1' s='{style_id}' t='inlineStr'><is><t>{value}</t></is></c></row></sheetData>"
            "<mergeCells count='1'><mergeCell ref='A1:B1'/></mergeCells>"
            "<pageMargins left='0.5' right='0.5' top='0.5' bottom='0.5' header='0.2' footer='0.2'/>"
            "</worksheet>"
        ),
    }
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for name, text in parts.items():
            archive.writestr(name, text)


def test_xlsx_guard_allows_text_and_blocks_layout_change(tmp_path):
    baseline = tmp_path / "a.xlsx"
    text_change = tmp_path / "b.xlsx"
    width_change = tmp_path / "c.xlsx"
    make_xlsx(baseline, "TEXTO A")
    make_xlsx(text_change, "TEXTO B")
    make_xlsx(width_change, "TEXTO A", width="13")
    assert compare_layout(baseline, text_change).match is True
    assert compare_layout(baseline, width_change).match is False


def make_patch_workbook(path: Path) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "CAPA"
    cover["N2"] = 1
    cover["P2"] = 2
    cover["A3"] = "TEXTO ORIGINAL"
    detail = wb.create_sheet("2")
    detail["AQ2"] = 2
    detail["AS2"] = 2
    detail["A3"] = "SERVIÇO ORIGINAL"
    wb.save(path)


def test_text_patch_changes_only_existing_cell_and_preserves_pagination(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    patch = tmp_path / "patch.json"
    make_patch_workbook(source)
    patch.write_text(json.dumps({
        "changes": [{"sheet": "2", "cell": "A3", "value": "SERVIÇO REVISADO"}],
        "update_pagination": True,
    }, ensure_ascii=False), encoding="utf-8")
    report = apply_patch(source, patch, output)
    wb = load_workbook(output, data_only=False)
    assert wb["2"]["A3"].value == "SERVIÇO REVISADO"
    assert wb["CAPA"]["P2"].value == 2
    assert wb["2"]["AS2"].value == 2
    assert report["approved_for_emission"] is True


def test_text_patch_rejects_cell_outside_matrix(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    patch = tmp_path / "patch.json"
    make_patch_workbook(source)
    patch.write_text(json.dumps({
        "changes": [{"sheet": "2", "cell": "ZZ999", "value": "X"}]
    }), encoding="utf-8")
    try:
        apply_patch(source, patch, output)
    except PatchError as exc:
        assert "fora da matriz" in str(exc)
    else:
        raise AssertionError("PatchError esperado")
    assert not output.exists()

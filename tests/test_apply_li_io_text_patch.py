import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pipeline.apply_li_io_text_patch import PatchError, apply_patch


def make_workbook(path: Path) -> None:
    wb = Workbook()
    cover = wb.active
    cover.title = "CAPA"
    cover["N2"] = 1
    cover["P2"] = 2
    cover["A3"] = "TEXTO ORIGINAL"
    cover.column_dimensions["A"].width = 20
    detail = wb.create_sheet("2")
    detail["AQ2"] = 2
    detail["AS2"] = 2
    detail["A3"] = "SERVIÇO ORIGINAL"
    detail.column_dimensions["A"].width = 20
    wb.save(path)


def test_text_patch_preserves_layout_and_updates_pagination(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    patch = tmp_path / "patch.json"
    make_workbook(source)
    patch.write_text(
        json.dumps({
            "changes": [{"sheet": "2", "cell": "A3", "value": "SERVIÇO REVISADO"}],
            "update_pagination": True,
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    report = apply_patch(source, patch, output)
    wb = load_workbook(output, data_only=False)
    assert wb["2"]["A3"].value == "SERVIÇO REVISADO"
    assert wb["CAPA"]["P2"].value == 2
    assert wb["2"]["AS2"].value == 2
    assert report["approved_for_emission"] is True
    assert report["critical_count"] == 0


def test_patch_rejects_cell_outside_existing_matrix(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    patch = tmp_path / "patch.json"
    make_workbook(source)
    patch.write_text(
        json.dumps({"changes": [{"sheet": "2", "cell": "ZZ999", "value": "X"}]}),
        encoding="utf-8",
    )
    try:
        apply_patch(source, patch, output)
    except PatchError as exc:
        assert "fora da matriz" in str(exc)
    else:
        raise AssertionError("PatchError esperado")
    assert not output.exists()
